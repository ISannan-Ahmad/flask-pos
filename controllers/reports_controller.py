import io
import csv
from datetime import datetime
from models import (Order, PurchaseOrder, Expense, EmployeePayment,
                     CustomerTransaction, PKT)
from extensions import db
from sqlalchemy import func, extract


class ReportsController:
    @staticmethod
    def get_monthly_report(year, month):
        """Gather all financial and sales data for a given month."""
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']

        # Sales
        sales_q = Order.query.filter(
            Order.status == 'approved',
            extract('year', Order.created_at) == year,
            extract('month', Order.created_at) == month)

        total_revenue = float(sales_q.with_entities(func.sum(Order.total_amount)).scalar() or 0)
        total_orders = sales_q.count()

        cash_q = sales_q.filter(Order.order_type == 'sale')
        credit_q = sales_q.filter(Order.order_type == 'credit_sale')

        cash_sales = float(cash_q.with_entities(func.sum(Order.total_amount)).scalar() or 0)
        credit_sales = float(credit_q.with_entities(func.sum(Order.total_amount)).scalar() or 0)

        # Purchases
        total_purchases = float(
            PurchaseOrder.query.filter(
                PurchaseOrder.status == 'received',
                extract('year', PurchaseOrder.created_at) == year,
                extract('month', PurchaseOrder.created_at) == month
            ).with_entities(func.sum(PurchaseOrder.total_amount)).scalar() or 0)

        # Expenses
        total_expenses = float(
            Expense.query.filter(
                extract('year', Expense.expense_date) == year,
                extract('month', Expense.expense_date) == month
            ).with_entities(func.sum(Expense.amount)).scalar() or 0)

        # Staff Expenses
        total_staff_expenses = float(
            EmployeePayment.query.filter(
                extract('year', EmployeePayment.date) == year,
                extract('month', EmployeePayment.date) == month
            ).with_entities(func.sum(EmployeePayment.amount)).scalar() or 0)

        # Receipts collected
        receipts_collected = float(
            CustomerTransaction.query.filter(
                CustomerTransaction.transaction_type == 'payment',
                extract('year', CustomerTransaction.created_at) == year,
                extract('month', CustomerTransaction.created_at) == month
            ).with_entities(func.sum(CustomerTransaction.amount)).scalar() or 0)

        outstanding_credit = credit_sales - receipts_collected
        net_profit = total_revenue - total_expenses - total_staff_expenses

        return {
            'year': year,
            'month': month,
            'month_name': month_names[month - 1],
            'total_revenue': total_revenue,
            'total_purchases': total_purchases,
            'total_expenses': total_expenses,
            'total_staff_expenses': total_staff_expenses,
            'net_profit': net_profit,
            'total_sales': total_revenue,
            'total_orders': total_orders,
            'cash_sales': cash_sales,
            'credit_sales': credit_sales,
            'receipts_collected': receipts_collected,
            'outstanding_credit': outstanding_credit,
        }

    @staticmethod
    def generate_csv(data):
        """Return CSV string for the monthly report."""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([f"Monthly Financial Report — {data['month_name']} {data['year']}"])
        writer.writerow([])

        writer.writerow(['FINANCIAL SUMMARY'])
        writer.writerow(['Metric', 'Amount (Rs.)'])
        writer.writerow(['Revenue', f"{data['total_revenue']:.2f}"])
        writer.writerow(['Purchases', f"{data['total_purchases']:.2f}"])
        writer.writerow(['Expenses', f"{data['total_expenses']:.2f}"])
        writer.writerow(['Staff Expenses', f"{data['total_staff_expenses']:.2f}"])
        writer.writerow(['Net Profit', f"{data['net_profit']:.2f}"])
        writer.writerow([])

        writer.writerow(['SALES METRICS'])
        writer.writerow(['Metric', 'Amount (Rs.)'])
        writer.writerow(['Total Sales', f"{data['total_sales']:.2f}"])
        writer.writerow(['Cash Sales', f"{data['cash_sales']:.2f}"])
        writer.writerow(['Credit Sales', f"{data['credit_sales']:.2f}"])
        writer.writerow(['Receipts Collected', f"{data['receipts_collected']:.2f}"])
        writer.writerow(['Outstanding Credit', f"{data['outstanding_credit']:.2f}"])
        writer.writerow(['Total Orders', data['total_orders']])

        return output.getvalue()

    @staticmethod
    def generate_excel(data):
        """Return BytesIO Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"{data['month_name']} {data['year']}"

        # Styles
        title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        currency_fmt = '#,##0.00'
        thin_border = Border(
            bottom=Side(style='thin', color='D9D9D9'))

        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Monthly Financial Report — {data['month_name']} {data['year']}"
        ws['A1'].font = title_font

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Financial Summary
        row = 3
        ws.cell(row=row, column=1, value='FINANCIAL SUMMARY').font = Font(bold=True, size=12)
        row += 1
        for col_idx, label in enumerate(['Metric', 'Amount (Rs.)'], 1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1
        fin_rows = [
            ('Revenue', data['total_revenue']),
            ('Purchases', data['total_purchases']),
            ('Expenses', data['total_expenses']),
            ('Staff Expenses', data['total_staff_expenses']),
            ('Net Profit', data['net_profit']),
        ]
        for label, value in fin_rows:
            ws.cell(row=row, column=1, value=label).border = thin_border
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = currency_fmt
            cell.border = thin_border
            if label == 'Net Profit':
                cell.font = Font(bold=True, color='006100' if value >= 0 else 'C00000')
            row += 1

        # Sales Metrics
        row += 1
        ws.cell(row=row, column=1, value='SALES METRICS').font = Font(bold=True, size=12)
        row += 1
        for col_idx, label in enumerate(['Metric', 'Amount (Rs.)'], 1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1
        sales_rows = [
            ('Total Sales', data['total_sales']),
            ('Cash Sales', data['cash_sales']),
            ('Credit Sales', data['credit_sales']),
            ('Receipts Collected', data['receipts_collected']),
            ('Outstanding Credit', data['outstanding_credit']),
        ]
        for label, value in sales_rows:
            ws.cell(row=row, column=1, value=label).border = thin_border
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = currency_fmt
            cell.border = thin_border
            row += 1

        ws.cell(row=row, column=1, value='Total Orders').border = thin_border
        ws.cell(row=row, column=2, value=data['total_orders']).border = thin_border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(data):
        """Return BytesIO PDF report."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Title'],
            fontSize=18, textColor=colors.HexColor('#1F4E79'),
            spaceAfter=12)
        section_style = ParagraphStyle(
            'SectionTitle', parent=styles['Heading2'],
            fontSize=13, textColor=colors.HexColor('#1F4E79'),
            spaceBefore=16, spaceAfter=8)

        elements = []

        # Title
        elements.append(Paragraph(
            f"Monthly Financial Report — {data['month_name']} {data['year']}", title_style))
        elements.append(Spacer(1, 6*mm))

        def fmt(v):
            return f"Rs. {v:,.2f}"

        # Financial Summary Table
        elements.append(Paragraph('Financial Summary', section_style))
        fin_data = [
            ['Metric', 'Amount'],
            ['Revenue', fmt(data['total_revenue'])],
            ['Purchases', fmt(data['total_purchases'])],
            ['Expenses', fmt(data['total_expenses'])],
            ['Staff Expenses', fmt(data['total_staff_expenses'])],
            ['Net Profit', fmt(data['net_profit'])],
        ]
        t = Table(fin_data, colWidths=[100*mm, 60*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 6*mm))

        # Sales Metrics Table
        elements.append(Paragraph('Sales Metrics', section_style))
        sales_data = [
            ['Metric', 'Amount'],
            ['Total Sales', fmt(data['total_sales'])],
            ['Cash Sales', fmt(data['cash_sales'])],
            ['Credit Sales', fmt(data['credit_sales'])],
            ['Receipts Collected', fmt(data['receipts_collected'])],
            ['Outstanding Credit', fmt(data['outstanding_credit'])],
            ['Total Orders', str(data['total_orders'])],
        ]
        t2 = Table(sales_data, colWidths=[100*mm, 60*mm])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ]))
        elements.append(t2)

        # Footer
        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph(
            f"Generated on {datetime.now(PKT).strftime('%Y-%m-%d %H:%M')} PKT",
            styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer
